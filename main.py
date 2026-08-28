import argparse
import csv
import hashlib
import json
import math
from pathlib import Path
from typing import Any


VERSION = "0.2"

WEIGHTS = {
    "price": 0.50,
    "lead_time": 0.30,
    "payment_terms": 0.20,
}

REQUIRED_FIELDS = [
    "name",
    "currency",
    "price",
    "lead_time_weeks",
    "payment_days",
]

RESERVED_FIELDS = {"rfqdiff_source"}


def _coerce_number(value: Any, field: str, source: str) -> int | float:
    if isinstance(value, bool):
        raise ValueError(f"{source}: {field} must be numeric")

    if isinstance(value, (int, float)):
        number = value
    else:
        try:
            number = float(str(value).strip())
        except (TypeError, ValueError) as error:
            raise ValueError(f"{source}: {field} must be numeric") from error

    if isinstance(number, float) and number.is_integer():
        return int(number)
    return number


def validate_quote(quote: dict[str, Any], source: str) -> dict[str, Any]:
    reserved = sorted(RESERVED_FIELDS.intersection(quote))
    if reserved:
        raise ValueError(
            f"{source}: reserved field(s) cannot be supplied: {', '.join(reserved)}"
        )

    missing = [field for field in REQUIRED_FIELDS if field not in quote]
    if missing:
        raise ValueError(
            f"{source}: missing required field(s): {', '.join(missing)}"
        )

    validated = quote.copy()
    validated["price"] = _coerce_number(validated["price"], "price", source)
    validated["lead_time_weeks"] = _coerce_number(
        validated["lead_time_weeks"], "lead_time_weeks", source
    )
    validated["payment_days"] = _coerce_number(
        validated["payment_days"], "payment_days", source
    )

    if validated["price"] <= 0:
        raise ValueError(f"{source}: price must be greater than 0")

    if validated["lead_time_weeks"] <= 0:
        raise ValueError(f"{source}: lead_time_weeks must be greater than 0")

    if validated["payment_days"] < 0:
        raise ValueError(f"{source}: payment_days cannot be negative")

    validated["currency"] = str(validated["currency"]).strip().upper()
    if not validated["currency"]:
        raise ValueError(f"{source}: currency cannot be empty")

    validated["name"] = str(validated["name"]).strip()
    if not validated["name"]:
        raise ValueError(f"{source}: name cannot be empty")

    return validated


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def attach_source(
    quote: dict[str, Any],
    path: Path,
    source_format: str,
    sha256: str,
    *,
    row: int | None = None,
    sheet: str | None = None,
) -> dict[str, Any]:
    sourced = quote.copy()
    source: dict[str, Any] = {
        "file": path.name,
        "format": source_format,
        "sha256": sha256,
    }
    if row is not None:
        source["row"] = row
    if sheet is not None:
        source["sheet"] = sheet
    sourced["rfqdiff_source"] = source
    return sourced


def load_quote(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as file:
        quote = json.load(file)

    if not isinstance(quote, dict):
        raise ValueError(f"{path}: quotation JSON must contain one object")

    validated = validate_quote(quote, str(path))
    return attach_source(validated, path, "json", file_sha256(path))


def load_csv_quotes(path: Path) -> list[dict[str, Any]]:
    sha256 = file_sha256(path)
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        fieldnames = reader.fieldnames or []
        missing = [field for field in REQUIRED_FIELDS if field not in fieldnames]
        if missing:
            raise ValueError(
                f"{path}: missing required column(s): {', '.join(missing)}"
            )

        quotes = []
        for row_number, row in enumerate(reader, start=2):
            if not any(value not in (None, "") for value in row.values()):
                continue
            validated = validate_quote(dict(row), f"{path}: row {row_number}")
            quotes.append(
                attach_source(
                    validated,
                    path,
                    "csv",
                    sha256,
                    row=row_number,
                )
            )

    if not quotes:
        raise ValueError(f"{path}: no supplier quotations found")
    return quotes


def load_xlsx_quotes(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError(
            "Excel import requires openpyxl. Install the project package dependencies first."
        ) from error

    sha256 = file_sha256(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as error:
            raise ValueError(f"{path}: workbook is empty") from error

        fieldnames = [
            str(value).strip() if value is not None else "" for value in header_row
        ]
        missing = [field for field in REQUIRED_FIELDS if field not in fieldnames]
        if missing:
            raise ValueError(
                f"{path}: missing required column(s): {', '.join(missing)}"
            )

        quotes = []
        for row_number, values in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            row = dict(zip(fieldnames, values))
            validated = validate_quote(row, f"{path}: row {row_number}")
            quotes.append(
                attach_source(
                    validated,
                    path,
                    "xlsx",
                    sha256,
                    row=row_number,
                    sheet=worksheet.title,
                )
            )
    finally:
        workbook.close()

    if not quotes:
        raise ValueError(f"{path}: no supplier quotations found")
    return quotes


def load_quotes(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return [load_quote(path)]
    if suffix == ".csv":
        return load_csv_quotes(path)
    if suffix == ".xlsx":
        return load_xlsx_quotes(path)
    raise ValueError(
        f"{path}: unsupported quotation format '{suffix or '<none>'}'. "
        "Use .json, .csv or .xlsx."
    )


def validate_currencies(quotes: list[dict[str, Any]]) -> str:
    currencies = {quote["currency"].upper() for quote in quotes}
    if len(currencies) != 1:
        raise ValueError(
            "All quotations must use the same currency. "
            "Normalize mixed-currency quotations first with currency-normalizer."
        )
    return next(iter(currencies))


def validate_weights(weights: dict[str, Any]) -> dict[str, float]:
    expected = set(WEIGHTS)
    supplied = set(weights)

    missing = sorted(expected - supplied)
    if missing:
        raise ValueError(f"weights: missing required key(s): {', '.join(missing)}")

    unsupported = sorted(supplied - expected)
    if unsupported:
        raise ValueError(f"weights: unsupported key(s): {', '.join(unsupported)}")

    validated: dict[str, float] = {}
    for key in WEIGHTS:
        value = weights[key]
        if isinstance(value, bool):
            raise ValueError(f"weights: {key} must be numeric")
        try:
            number = float(value)
        except (TypeError, ValueError) as error:
            raise ValueError(f"weights: {key} must be numeric") from error
        if number < 0 or number > 1:
            raise ValueError(f"weights: {key} must be between 0 and 1")
        validated[key] = number

    total = sum(validated.values())
    if not math.isclose(total, 1.0, rel_tol=0.0, abs_tol=1e-9):
        raise ValueError(f"weights: values must sum to 1.0; got {total:g}")

    return validated


def load_weights(path: Path) -> dict[str, float]:
    with path.open("r", encoding="utf-8") as file:
        weights = json.load(file)

    if not isinstance(weights, dict):
        raise ValueError(f"{path}: weights JSON must contain one object")

    try:
        return validate_weights(weights)
    except ValueError as error:
        raise ValueError(f"{path}: {error}") from error


def score_quotes(
    quotes: list[dict[str, Any]],
    weights: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    effective_weights = validate_weights(WEIGHTS if weights is None else weights)
    min_price = min(quote["price"] for quote in quotes)
    min_lead_time = min(quote["lead_time_weeks"] for quote in quotes)
    max_payment_days = max(quote["payment_days"] for quote in quotes)

    scored_quotes = []

    for quote in quotes:
        price_score = (
            min_price / quote["price"]
        ) * effective_weights["price"] * 100
        lead_time_score = (
            min_lead_time / quote["lead_time_weeks"]
        ) * effective_weights["lead_time"] * 100

        if max_payment_days == 0:
            payment_score = effective_weights["payment_terms"] * 100
        else:
            payment_score = (
                quote["payment_days"] / max_payment_days
            ) * effective_weights["payment_terms"] * 100

        total_score = price_score + lead_time_score + payment_score

        scored_quote = quote.copy()
        scored_quote["score"] = round(total_score, 1)
        scored_quotes.append(scored_quote)

    return sorted(
        scored_quotes,
        key=lambda item: (-item["score"], item["name"].lower()),
    )


def build_result(
    scored_quotes: list[dict[str, Any]],
    currency: str,
    weights: dict[str, Any] | None = None,
) -> dict[str, Any]:
    effective_weights = validate_weights(WEIGHTS if weights is None else weights)
    winner = scored_quotes[0]
    cheapest = min(scored_quotes, key=lambda item: item["price"])
    fastest = min(scored_quotes, key=lambda item: item["lead_time_weeks"])
    best_terms = max(scored_quotes, key=lambda item: item["payment_days"])

    return {
        "tool": "rfqdiff",
        "version": VERSION,
        "currency": currency,
        "recommended_supplier": winner["name"],
        "suppliers": scored_quotes,
        "decision_summary": {
            "recommended_supplier": {
                "name": winner["name"],
                "score": winner["score"],
            },
            "lowest_price": {
                "name": cheapest["name"],
                "price": cheapest["price"],
            },
            "fastest_lead_time": {
                "name": fastest["name"],
                "lead_time_weeks": fastest["lead_time_weeks"],
            },
            "best_payment_terms": {
                "name": best_terms["name"],
                "payment_days": best_terms["payment_days"],
            },
        },
        "weights": effective_weights,
    }


def money(value: float, currency: str) -> str:
    return f"{currency} {value:,.2f}"


def print_report(
    quotes: list[dict[str, Any]],
    currency: str,
    weights: dict[str, Any] | None = None,
) -> None:
    effective_weights = validate_weights(WEIGHTS if weights is None else weights)
    print(f"\nRFQDIFF v{VERSION}")
    print("=" * 86)
    print(
        f"{'Supplier':<24}"
        f"{'Price':>18}"
        f"{'Lead time':>14}"
        f"{'Payment':>14}"
        f"{'Score':>12}"
    )
    print("-" * 86)

    for quote in quotes:
        print(
            f"{quote['name']:<24}"
            f"{money(quote['price'], currency):>18}"
            f"{str(quote['lead_time_weeks']) + ' weeks':>14}"
            f"{str(quote['payment_days']) + ' days':>14}"
            f"{str(quote['score']) + '/100':>12}"
        )

    payload = build_result(quotes, currency, effective_weights)
    summary = payload["decision_summary"]
    winner = summary["recommended_supplier"]
    cheapest = summary["lowest_price"]
    fastest = summary["fastest_lead_time"]
    best_terms = summary["best_payment_terms"]

    print("\nDecision summary")
    print("-" * 86)
    print(f"Recommended supplier : {winner['name']} ({winner['score']}/100)")
    print(
        f"Lowest price         : {cheapest['name']} "
        f"({money(cheapest['price'], currency)})"
    )
    print(
        f"Fastest lead time    : {fastest['name']} "
        f"({fastest['lead_time_weeks']} weeks)"
    )
    print(
        f"Best payment terms   : {best_terms['name']} "
        f"({best_terms['payment_days']} days)"
    )

    print("\nScoring weights")
    print(
        f"Price {effective_weights['price']:.0%} | "
        f"Lead time {effective_weights['lead_time']:.0%} | "
        f"Payment terms {effective_weights['payment_terms']:.0%}\n"
        "Lower price and lead time score higher; longer payment terms score higher."
    )


def write_json(payload: dict[str, Any], path: Path) -> None:
    path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def _report_source_fields(quote: dict[str, Any]) -> dict[str, Any]:
    source = quote.get("rfqdiff_source", {})
    return {
        "source_file": source.get("file", ""),
        "source_format": source.get("format", ""),
        "source_sha256": source.get("sha256", ""),
        "source_row": source.get("row", ""),
        "source_sheet": source.get("sheet", ""),
    }


def write_csv_report(payload: dict[str, Any], path: Path) -> None:
    fieldnames = [
        "rank",
        "recommended",
        "name",
        "currency",
        "price",
        "lead_time_weeks",
        "payment_days",
        "score",
        "source_file",
        "source_format",
        "source_sha256",
        "source_row",
        "source_sheet",
    ]
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for rank, quote in enumerate(payload["suppliers"], start=1):
            row = {
                "rank": rank,
                "recommended": quote["name"] == payload["recommended_supplier"],
                "name": quote["name"],
                "currency": payload["currency"],
                "price": quote["price"],
                "lead_time_weeks": quote["lead_time_weeks"],
                "payment_days": quote["payment_days"],
                "score": quote["score"],
            }
            row.update(_report_source_fields(quote))
            writer.writerow(row)


def write_xlsx_report(payload: dict[str, Any], path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as error:
        raise ValueError(
            "Excel report export requires openpyxl. Install the project package dependencies first."
        ) from error

    workbook = Workbook()
    comparison = workbook.active
    comparison.title = "Comparison"
    comparison.append(
        [
            "Rank",
            "Recommended",
            "Supplier",
            "Currency",
            "Price",
            "Lead Time (Weeks)",
            "Payment Days",
            "Score",
            "Source File",
            "Source Format",
            "Source SHA-256",
            "Source Row",
            "Source Sheet",
        ]
    )
    for rank, quote in enumerate(payload["suppliers"], start=1):
        source = quote.get("rfqdiff_source", {})
        comparison.append(
            [
                rank,
                quote["name"] == payload["recommended_supplier"],
                quote["name"],
                payload["currency"],
                quote["price"],
                quote["lead_time_weeks"],
                quote["payment_days"],
                quote["score"],
                source.get("file", ""),
                source.get("format", ""),
                source.get("sha256", ""),
                source.get("row", ""),
                source.get("sheet", ""),
            ]
        )
    comparison.freeze_panes = "A2"
    comparison.auto_filter.ref = comparison.dimensions

    summary = workbook.create_sheet("Summary")
    decision = payload["decision_summary"]
    summary.append(["Metric", "Supplier", "Value"])
    summary.append(
        [
            "Recommended supplier",
            decision["recommended_supplier"]["name"],
            decision["recommended_supplier"]["score"],
        ]
    )
    summary.append(
        [
            "Lowest price",
            decision["lowest_price"]["name"],
            decision["lowest_price"]["price"],
        ]
    )
    summary.append(
        [
            "Fastest lead time",
            decision["fastest_lead_time"]["name"],
            decision["fastest_lead_time"]["lead_time_weeks"],
        ]
    )
    summary.append(
        [
            "Best payment terms",
            decision["best_payment_terms"]["name"],
            decision["best_payment_terms"]["payment_days"],
        ]
    )
    summary.append([])
    summary.append(["Scoring criterion", "Weight"])
    for criterion, weight in payload["weights"].items():
        summary.append([criterion, weight])
    summary.freeze_panes = "A2"

    workbook.save(path)
    workbook.close()


def write_report(payload: dict[str, Any], path: Path) -> None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        write_csv_report(payload, path)
        return
    if suffix == ".xlsx":
        write_xlsx_report(payload, path)
        return
    raise ValueError(
        f"{path}: unsupported report format '{suffix or '<none>'}'. Use .csv or .xlsx."
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare supplier quotations from JSON, CSV or Excel files."
    )
    parser.add_argument(
        "quotes",
        nargs="+",
        type=Path,
        help="Quotation files (.json, .csv or .xlsx). Tabular files may contain multiple suppliers.",
    )
    parser.add_argument(
        "--weights",
        type=Path,
        help="JSON file with price, lead_time and payment_terms weights summing to 1.0.",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Return a machine-readable comparison payload.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Write the machine-readable comparison payload to a JSON file.",
    )
    parser.add_argument(
        "--report",
        type=Path,
        help="Write a comparison report as .csv or .xlsx.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    try:
        quotes = [quote for path in args.quotes for quote in load_quotes(path)]
        if len(quotes) < 2:
            raise ValueError("rfqdiff needs at least two supplier quotations.")
        currency = validate_currencies(quotes)
        weights = load_weights(args.weights) if args.weights else validate_weights(WEIGHTS)
        scored_quotes = score_quotes(quotes, weights)
        payload = build_result(scored_quotes, currency, weights)

        if args.output:
            write_json(payload, args.output)
        if args.report:
            write_report(payload, args.report)
    except (OSError, json.JSONDecodeError, ValueError) as error:
        raise SystemExit(f"Error: {error}") from error

    if args.json:
        print(json.dumps(payload, indent=2, ensure_ascii=False))
    else:
        print_report(scored_quotes, currency, weights)


if __name__ == "__main__":
    main()
