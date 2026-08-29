import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

import main as rfqdiff


PILOT_FIXTURE = (
    Path(__file__).resolve().parents[1] / "samples" / "pilot" / "industrial-rfq.csv"
)
PILOT_SHA256 = "c0627382a79415ec0daf90335d28d9a1b906366aae60e90b5c463d5f02982ed8"


class PilotGoldenIntegrationTests(unittest.TestCase):
    def test_sanitized_industrial_rfq_pipeline_is_stable(self) -> None:
        quotes = rfqdiff.load_quotes(PILOT_FIXTURE)
        currency = rfqdiff.validate_currencies(quotes)
        scored = rfqdiff.score_quotes(quotes)
        payload = rfqdiff.build_result(scored, currency)

        self.assertEqual(currency, "EUR")
        self.assertEqual(
            [supplier["name"] for supplier in scored],
            [
                "Crest Manufacturing",
                "Alpha Components",
                "Bravo Industrial",
                "Delta Engineering",
            ],
        )
        self.assertEqual(payload["recommended_supplier"], "Crest Manufacturing")
        self.assertEqual(payload["decision_summary"]["runner_up"], {"name": "Alpha Components", "score": 80.8})
        self.assertEqual(payload["decision_summary"]["score_margin"], 12.7)
        self.assertEqual(
            payload["decision_explanation"]["criterion_leaders"],
            {
                "price": "Delta Engineering",
                "lead_time": "Crest Manufacturing",
                "payment_terms": "Crest Manufacturing",
            },
        )
        self.assertEqual(
            payload["decision_explanation"]["winner_score_breakdown"],
            {"price": 43.5, "lead_time": 30.0, "payment_terms": 20.0},
        )

        for row_number, quote in enumerate(quotes, start=2):
            self.assertEqual(quote["rfqdiff_source"]["file"], "industrial-rfq.csv")
            self.assertEqual(quote["rfqdiff_source"]["format"], "csv")
            self.assertEqual(quote["rfqdiff_source"]["sha256"], PILOT_SHA256)
            self.assertEqual(quote["rfqdiff_source"]["row"], row_number)

    def test_pilot_payload_exports_reviewable_csv_and_xlsx_reports(self) -> None:
        quotes = rfqdiff.load_quotes(PILOT_FIXTURE)
        currency = rfqdiff.validate_currencies(quotes)
        payload = rfqdiff.build_result(rfqdiff.score_quotes(quotes), currency)

        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            csv_path = directory_path / "comparison.csv"
            xlsx_path = directory_path / "comparison.xlsx"
            rfqdiff.write_report(payload, csv_path)
            rfqdiff.write_report(payload, xlsx_path)

            with csv_path.open("r", encoding="utf-8", newline="") as file:
                rows = list(csv.DictReader(file))

            self.assertEqual(rows[0]["name"], "Crest Manufacturing")
            self.assertEqual(rows[0]["recommended"], "True")
            self.assertEqual(rows[0]["score"], "93.5")
            self.assertEqual(rows[0]["source_file"], "industrial-rfq.csv")
            self.assertEqual(rows[0]["source_sha256"], PILOT_SHA256)

            workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Comparison", "Summary"])
                comparison = workbook["Comparison"]
                summary = workbook["Summary"]
                self.assertEqual(comparison["C2"].value, "Crest Manufacturing")
                self.assertEqual(comparison["B2"].value, True)
                self.assertEqual(comparison["H2"].value, 93.5)
                self.assertEqual(summary["B2"].value, "Crest Manufacturing")
                self.assertEqual(summary["B3"].value, "Alpha Components")
                self.assertEqual(summary["C4"].value, 12.7)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
