import csv
import tempfile
import unittest
from pathlib import Path

import main as rfqdiff


class ReportExportTests(unittest.TestCase):
    def setUp(self) -> None:
        quotes = [
            {
                "name": "Supplier A",
                "currency": "EUR",
                "price": 84200,
                "lead_time_weeks": 8,
                "payment_days": 30,
            },
            {
                "name": "Supplier B",
                "currency": "EUR",
                "price": 79400,
                "lead_time_weeks": 14,
                "payment_days": 0,
            },
        ]
        self.scored = rfqdiff.score_quotes(quotes)
        self.payload = rfqdiff.build_result(self.scored, "EUR")

    def test_write_csv_report_exports_ranked_suppliers(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "comparison.csv"
            rfqdiff.write_report(self.payload, path)

            with path.open("r", encoding="utf-8", newline="") as file:
                rows = list(csv.DictReader(file))

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["rank"], "1")
        self.assertEqual(rows[0]["name"], "Supplier A")
        self.assertEqual(rows[0]["recommended"], "True")
        self.assertEqual(rows[0]["currency"], "EUR")
        self.assertEqual(rows[0]["price_score"], "47.1")
        self.assertEqual(rows[0]["lead_time_score"], "30.0")
        self.assertEqual(rows[0]["payment_terms_score"], "20.0")

    def test_write_xlsx_report_exports_comparison_and_summary(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "comparison.xlsx"
            rfqdiff.write_report(self.payload, path)

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Comparison", "Summary"])
                comparison = workbook["Comparison"]
                self.assertEqual(comparison["C2"].value, "Supplier A")
                self.assertTrue(comparison["B2"].value)
                self.assertEqual(comparison["I2"].value, 47.1)
                self.assertEqual(comparison["J2"].value, 30.0)
                self.assertEqual(comparison["K2"].value, 20.0)

                summary = workbook["Summary"]
                self.assertEqual(summary["A2"].value, "Recommended supplier")
                self.assertEqual(summary["B2"].value, "Supplier A")
                self.assertEqual(summary["A3"].value, "Runner-up")
                self.assertEqual(summary["B3"].value, "Supplier B")
                self.assertEqual(summary["A4"].value, "Score margin")
                self.assertEqual(summary["C4"].value, 10.0)
                self.assertEqual(summary["A10"].value, "price")
                self.assertEqual(summary["B10"].value, 0.5)
            finally:
                workbook.close()

    def test_write_report_rejects_unsupported_extension(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "comparison.pdf"
            with self.assertRaisesRegex(ValueError, "unsupported report format"):
                rfqdiff.write_report(self.payload, path)


if __name__ == "__main__":
    unittest.main()
